# import các module openpyxl và tkinter
from openpyxl import *
from tkinter import *

# khai báo biến wb và sheet toàn cầu

# mở tệp excel hiện có
wb = load_workbook('Lab4//excel.xlsx')

# tạo đối tượng sheet
sheet = wb.active


def excel():
    
    # điều chỉnh độ rộng của cột trong bảng tính excel
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # ghi dữ liệu đã cho vào bảng tính excel
    # tại vị trí cụ thể
    sheet.cell(row=1, column=1).value = "Tên"
    sheet.cell(row=1, column=2).value = "Khóa học"
    sheet.cell(row=1, column=3).value = "Học kỳ"
    sheet.cell(row=1, column=4).value = "Số biểu mẫu"
    sheet.cell(row=1, column=5).value = "Số liên lạc"
    sheet.cell(row=1, column=6).value = "Email"
    sheet.cell(row=1, column=7).value = "Địa chỉ"


# Hàm để thiết lập trọng tâm (con trỏ)
def focus1(event):
    # đặt trọng tâm vào hộp course_field
    course_field.focus_set()


# Hàm để thiết lập trọng tâm
def focus2(event):
    # đặt trọng tâm vào hộp sem_field
    sem_field.focus_set()


# Hàm để thiết lập trọng tâm
def focus3(event):
    # đặt trọng tâm vào hộp form_no_field
    form_no_field.focus_set()


# Hàm để thiết lập trọng tâm
def focus4(event):
    # đặt trọng tâm vào hộp contact_no_field
    contact_no_field.focus_set()


# Hàm để thiết lập trọng tâm
def focus5(event):
    # đặt trọng tâm vào hộp email_id_field
    email_id_field.focus_set()


# Hàm để thiết lập trọng tâm
def focus6(event):
    # đặt trọng tâm vào hộp address_field
    address_field.focus_set()


# Hàm để xóa nội dung của các ô nhập văn bản
def clear():
    
    # xóa nội dung của ô nhập văn bản
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)


# Hàm để lấy dữ liệu từ cửa sổ GUI
# và ghi vào tệp excel
def insert():
    
    # nếu người dùng không điền vào bất kỳ ô nào
    # thì in ra "empty input"
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
            
        print("empty input")

    else:

        # gán giá trị hàng tối đa và cột tối đa
        # mà dữ liệu được viết vào
        # trong bảng tính excel cho biến
        current_row = sheet.max_row
        current_column = sheet.max_column

        # phương thức get trả về văn bản hiện tại
        # dưới dạng chuỗi mà chúng ta viết vào
        # bảng tính excel tại vị trí cụ thể
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = course_field.get()
        sheet.cell(row=current_row + 1, column=3).value = sem_field.get()
        sheet.cell(row=current_row + 1, column=4).value = form_no_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()

        # lưu tệp
        wb.save('Lab4//excel.xlsx')

        # đặt trọng tâm vào hộp name_field
        name_field.focus_set()

        # gọi hàm clear()
        clear()


# Mã điều khiển chính
if __name__ == "__main__":
    
    # tạo một cửa sổ GUI
    root = Tk()

    # đặt màu nền của cửa sổ GUI
    root.configure(background='light green')

    # đặt tiêu đề của cửa sổ GUI
    root.title("Form đăng ký")

    # đặt cấu hình của cửa sổ GUI
    root.geometry("500x300")

    excel()

    # tạo một nhãn Form
    heading = Label(root, text="Form", bg="light green")

    # tạo một nhãn Name
    name = Label(root, text="Tên", bg="light green")

    # tạo một nhãn Course
    course = Label(root, text="Khóa học", bg="light green")

    # tạo một nhãn Semester
    sem = Label(root, text="Học kỳ", bg="light green")

    # tạo một nhãn Form No.
    form_no = Label(root, text="Số biểu mẫu", bg="light green")

    # tạo một nhãn Contact No.
    contact_no = Label(root, text="Số liên lạc", bg="light green")

    # tạo một nhãn Email id
    email_id = Label(root, text="Email", bg="light green")

    # tạo một nhãn Địa chỉ
    address = Label(root, text="Địa chỉ", bg="light green")

    # phương thức lưới được sử dụng để đặt
    # các widget ở vị trí tương ứng
    # trong cấu trúc dạng bảng.
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    course.grid(row=2, column=0)
    sem.grid(row=3, column=0)
    form_no.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)

    # tạo một hộp nhập văn bản
    # để nhập thông tin
    name_field = Entry(root)
    course_field = Entry(root)
    sem_field = Entry(root)
    form_no_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)

    # phương thức bind của widget được sử dụng để
    # gắn hàm với các sự kiện

    # mỗi khi phím enter được nhấn
    # sau đó gọi hàm focus1
    name_field.bind("<Return>", focus1)

    # mỗi khi phím enter được nhấn
    # sau đó gọi hàm focus2
    course_field.bind("<Return>", focus2)

    # mỗi khi phím enter được nhấn
    # sau đó gọi hàm focus3
    sem_field.bind("<Return>", focus3)

    # mỗi khi phím enter được nhấn
    # sau đó gọi hàm focus4
    form_no_field.bind("<Return>", focus4)

    # mỗi khi phím enter được nhấn
    # sau đó gọi hàm focus5
    contact_no_field.bind("<Return>", focus5)

    # mỗi khi phím enter được nhấn
    # sau đó gọi hàm focus6
    email_id_field.bind("<Return>", focus6)

    # phương thức lưới được sử dụng để đặt
    # các widget ở vị trí tương ứng
    # trong cấu trúc dạng bảng.
    name_field.grid(row=1, column=1, ipadx="100")
    course_field.grid(row=2, column=1, ipadx="100")
    sem_field.grid(row=3, column=1, ipadx="100")
    form_no_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")

    # gọi hàm excel
    excel()

    # tạo một nút Submit và đặt vào cửa sổ root
    submit = Button(root, text="Gửi", fg="Black",
                            bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # bắt đầu GUI
    root.mainloop()
